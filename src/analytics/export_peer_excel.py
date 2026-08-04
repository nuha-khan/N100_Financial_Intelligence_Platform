import sqlite3
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill,
    Font,
    Border,
    Side,
    Alignment,
)
from openpyxl.utils import get_column_letter

DB_PATH = "data/nifty100.db"
OUTPUT_PATH = "outputs/peer_comparison.xlsx"

# ------------------------------------------------------------------
# Metrics to include in Excel
# ------------------------------------------------------------------

METRICS = [
    "return_on_equity_pct",
    "return_on_capital_employed_pct",
    "net_profit_margin_pct",
    "debt_to_equity",
    "interest_coverage",
    "asset_turnover",
    "free_cash_flow_cr",
    "revenue_cagr_5y",
    "pat_cagr_5y",
    "eps_cagr_5y",
]

DISPLAY_NAMES = {
    "return_on_equity_pct": "ROE (%)",
    "return_on_capital_employed_pct": "ROCE (%)",
    "net_profit_margin_pct": "Net Profit Margin (%)",
    "debt_to_equity": "Debt / Equity",
    "interest_coverage": "Interest Coverage",
    "asset_turnover": "Asset Turnover",
    "free_cash_flow_cr": "Free Cash Flow",
    "revenue_cagr_5y": "Revenue CAGR (5Y)",
    "pat_cagr_5y": "PAT CAGR (5Y)",
    "eps_cagr_5y": "EPS CAGR (5Y)",
}


# ------------------------------------------------------------------
# Load Tables
# ------------------------------------------------------------------

def load_tables():

    conn = sqlite3.connect(DB_PATH)

    companies = pd.read_sql(
        "SELECT * FROM companies",
        conn,
    )

    financial_ratios = pd.read_sql(
        "SELECT * FROM financial_ratios",
        conn,
    )

    peer_groups = pd.read_sql(
        "SELECT * FROM peer_groups",
        conn,
    )

    peer_percentiles = pd.read_sql(
        "SELECT * FROM peer_percentiles",
        conn,
    )

    conn.close()

    return (
        companies,
        financial_ratios,
        peer_groups,
        peer_percentiles,
    )


# ------------------------------------------------------------------
# Latest Financial Year
# ------------------------------------------------------------------

def latest_financials(financial_ratios):

    latest_year = financial_ratios["year"].max()

    latest = financial_ratios[
        financial_ratios["year"] == latest_year
    ].copy()

    return latest


# ------------------------------------------------------------------
# Pivot percentile table
# ------------------------------------------------------------------

def pivot_percentiles(peer_percentiles):

    latest_year = peer_percentiles["year"].max()

    latest = peer_percentiles[
        peer_percentiles["year"] == latest_year
    ]

    pivot = latest.pivot_table(
        index="company_id",
        columns="metric",
        values="percentile_rank",
    )

    pivot = pivot.reset_index()

    rename_columns = {
        metric: f"{DISPLAY_NAMES[metric]} Percentile"
        for metric in METRICS
    }

    pivot.rename(
    columns=rename_columns,
    inplace=True,)

    # Remove the pivoted columns name
    pivot.columns.name = None

    # Ensure company_id is a normal column
    pivot = pivot.reset_index(drop=True)

    return pivot

# ------------------------------------------------------------------
# Merge Everything
# ------------------------------------------------------------------

def build_master_dataframe():

    (
        companies,
        financial_ratios,
        peer_groups,
        peer_percentiles,
    ) = load_tables()

    latest = latest_financials(
        financial_ratios,
    )

    percentile_df = pivot_percentiles(
        peer_percentiles,
    )

    # Rename id -> company_id for merge
    companies = companies.rename(
        columns={
            "id": "company_id",
        }
    )

    master = (
        latest
        .merge(
            companies[
                [
                    "company_id",
                    "company_name",
                ]
            ],
            on="company_id",
            how="left",
        )
        .merge(
            peer_groups,
            on="company_id",
            how="left",
        )
        .merge(
            percentile_df,
            on="company_id",
            how="left",
        )
    )

    return master

# ------------------------------------------------------------------
# Workbook Styles
# ------------------------------------------------------------------

HEADER_FILL = PatternFill(
    fill_type="solid",
    start_color="1F4E78",
    end_color="1F4E78",
)

HEADER_FONT = Font(
    color="FFFFFF",
    bold=True,
)

GREEN_FILL = PatternFill(
    fill_type="solid",
    start_color="C6EFCE",
    end_color="C6EFCE",
)

YELLOW_FILL = PatternFill(
    fill_type="solid",
    start_color="FFEB9C",
    end_color="FFEB9C",
)

RED_FILL = PatternFill(
    fill_type="solid",
    start_color="FFC7CE",
    end_color="FFC7CE",
)

BENCHMARK_FILL = PatternFill(
    fill_type="solid",
    start_color="FFD966",
    end_color="FFD966",
)

MEDIAN_FILL = PatternFill(
    fill_type="solid",
    start_color="D9EAD3",
    end_color="D9EAD3",
)

THIN_BORDER = Border(

    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),

)

CENTER = Alignment(
    horizontal="center",
    vertical="center",
)


# ------------------------------------------------------------------
# Build Workbook
# ------------------------------------------------------------------

def create_peer_workbook(master_df):

    workbook = Workbook()

    workbook.remove(workbook.active)

    peer_groups = sorted(
        master_df["peer_group_name"]
        .dropna()
        .unique()
    )

    for peer_group in peer_groups:

        sheet = workbook.create_sheet(
            title=peer_group[:31]
        )

        group = (
            master_df[
                master_df["peer_group_name"] == peer_group
            ]
            .copy()
            .sort_values(
                "composite_quality_score",
                ascending=False,
            )
        )

        headers = [
            "Company ID",
            "Company Name",
        ]

        for metric in METRICS:
            headers.append(DISPLAY_NAMES[metric])

        for metric in METRICS:
            headers.append(
                DISPLAY_NAMES[metric] + " Percentile"
            )

        headers.append("Benchmark")

        sheet.append(headers)

        for cell in sheet[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = THIN_BORDER
            cell.alignment = CENTER

        for _, row in group.iterrows():

            values = [

                row["company_id"],

                row["company_name"]
                if pd.notna(row["company_name"])
                else row["company_id"],

            ]

            for metric in METRICS:
                values.append(
                    row.get(metric)
                )

            for metric in METRICS:

                values.append(
                    row.get(
                        DISPLAY_NAMES[metric]
                        + " Percentile"
                    )
                )

            values.append(
                "Yes"
                if row["is_benchmark"] == 1
                else "No"
            )

            sheet.append(values)

        # ---------------------------
        # Median Row
        # ---------------------------

        median = [
            "",
            "Peer Median",
        ]

        for metric in METRICS:

            median.append(
                round(
                    group[metric].median(),
                    2,
                )
            )

        for metric in METRICS:

            percentile_column = (
                DISPLAY_NAMES[metric]
                + " Percentile"
            )

            median.append(
                round(
                    group[
                        percentile_column
                    ].median(),
                    2,
                )
            )

        median.append("")

        sheet.append(median)

    return workbook

# ------------------------------------------------------------------
# Apply Formatting
# ------------------------------------------------------------------

def format_workbook(workbook):

    for sheet in workbook.worksheets:

        max_row = sheet.max_row
        max_col = sheet.max_column

        # -----------------------------
        # Data rows
        # -----------------------------

        for row in range(2, max_row):

            benchmark = sheet.cell(
                row=row,
                column=max_col,
            ).value

            # Highlight benchmark company

            if benchmark == "Yes":

                for cell in sheet[row]:
                    cell.fill = BENCHMARK_FILL

            percentile_start = (
                2
                + len(METRICS)
                + 1
            )

            percentile_end = (
                percentile_start
                + len(METRICS)
                - 1
            )

            for col in range(
                percentile_start,
                percentile_end + 1,
            ):

                cell = sheet.cell(
                    row=row,
                    column=col,
                )

                value = cell.value

                if value is None:
                    continue

                if value >= 0.75:
                    cell.fill = GREEN_FILL

                elif value <= 0.25:
                    cell.fill = RED_FILL

                else:
                    cell.fill = YELLOW_FILL

        # -----------------------------
        # Median row
        # -----------------------------

        for cell in sheet[max_row]:

            cell.fill = MEDIAN_FILL
            cell.font = Font(bold=True)

        # -----------------------------
        # Borders
        # -----------------------------

        for row in sheet.iter_rows():

            for cell in row:

                cell.border = THIN_BORDER
                cell.alignment = CENTER

        # -----------------------------
        # Auto width
        # -----------------------------

        for column in sheet.columns:

            max_length = 0

            column_letter = get_column_letter(
                column[0].column
            )

            for cell in column:

                try:

                    max_length = max(
                        max_length,
                        len(str(cell.value)),
                    )

                except Exception:
                    pass

            sheet.column_dimensions[
                column_letter
            ].width = min(
                max_length + 3,
                35,
            )

        sheet.freeze_panes = "A2"

        sheet.auto_filter.ref = sheet.dimensions


# ------------------------------------------------------------------
# Export Workbook
# ------------------------------------------------------------------

def export_peer_workbook():

    master_df = build_master_dataframe()

    workbook = create_peer_workbook(
        master_df,
    )

    format_workbook(
        workbook,
    )

    output = Path(OUTPUT_PATH)

    output.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    workbook.save(output)

    print(
        f"\nWorkbook saved successfully → {output}"
    )


# ------------------------------------------------------------------
# Main
# ------------------------------------------------------------------

if __name__ == "__main__":

    export_peer_workbook()