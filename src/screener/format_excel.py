from openpyxl import load_workbook
from openpyxl.styles import PatternFill


GREEN = PatternFill(
    fill_type="solid",
    start_color="C6EFCE",
)

RED = PatternFill(
    fill_type="solid",
    start_color="FFC7CE",
)


def colour_screeners(path):

    wb = load_workbook(path)

    metric_columns = {
        "return_on_equity_pct",
        "debt_to_equity",
        "free_cash_flow_cr",
        "revenue_cagr_5y",
        "pat_cagr_5y",
        "pe_ratio",
        "pb_ratio",
        "dividend_yield_pct",
        "dividend_payout_ratio_pct",
        "market_cap_crore",
    }

    for sheet in wb.worksheets:

        headers = [
            c.value
            for c in sheet[1]
        ]

        preset = sheet.title.lower()

        for row in sheet.iter_rows(min_row=2):

            for cell in row:

                column = headers[cell.column - 1]

                if column not in metric_columns:
                    continue

                value = cell.value

                if value is None:
                    continue

                if preset == "quality_compounder":

                    if column == "return_on_equity_pct":
                        cell.fill = GREEN if value >= 15 else RED

                    elif column == "debt_to_equity":
                        cell.fill = GREEN if value <= 1 else RED

                    elif column == "free_cash_flow_cr":
                        cell.fill = GREEN if value > 0 else RED

                    elif column == "revenue_cagr_5y":
                        cell.fill = GREEN if value >= 10 else RED

                elif preset == "value_pick":

                    if column == "pe_ratio":
                        cell.fill = GREEN if value <= 20 else RED

                    elif column == "pb_ratio":
                        cell.fill = GREEN if value <= 3 else RED

                    elif column == "debt_to_equity":
                        cell.fill = GREEN if value <= 2 else RED

                    elif column == "dividend_yield_pct":
                        cell.fill = GREEN if value >= 1 else RED

                elif preset == "growth_accelerator":

                    if column == "pat_cagr_5y":
                        cell.fill = GREEN if value >= 20 else RED

                    elif column == "revenue_cagr_5y":
                        cell.fill = GREEN if value >= 15 else RED

                    elif column == "debt_to_equity":
                        cell.fill = GREEN if value <= 2 else RED

                elif preset == "dividend_champion":

                    if column == "dividend_yield_pct":
                        cell.fill = GREEN if value >= 2 else RED

                    elif column == "dividend_payout_ratio_pct":
                        cell.fill = GREEN if value <= 80 else RED

                    elif column == "free_cash_flow_cr":
                        cell.fill = GREEN if value > 0 else RED

                elif preset == "debt_free_blue_chip":

                    if column == "debt_to_equity":
                        cell.fill = GREEN if value == 0 else RED

                    elif column == "return_on_equity_pct":
                        cell.fill = GREEN if value >= 12 else RED

                    elif column == "market_cap_crore":
                        cell.fill = GREEN if value >= 5000 else RED

                elif preset == "turnaround_watch":

                    if column == "free_cash_flow_cr":
                        cell.fill = GREEN if value > 0 else RED

    wb.save(path)

    print("Excel formatting applied.")